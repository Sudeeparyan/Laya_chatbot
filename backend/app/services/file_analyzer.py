"""Pre-conversion file analysis to detect images and recommend conversion mode."""
from __future__ import annotations

from pathlib import Path

from app.models import FileAnalysisResponse

EXCEL_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm"}
PDF_EXTENSION = ".pdf"
OFFICE_EXTENSIONS = {".docx", ".doc", ".pptx"}


def analyze_file(file_path: Path, original_filename: str) -> FileAnalysisResponse:
    """Analyze a file and return recommendation for conversion mode."""
    extension = file_path.suffix.lower()
    file_size = file_path.stat().st_size

    has_images = False
    image_count = 0
    has_merged_cells = False
    sheet_count: int | None = None

    if extension in EXCEL_EXTENSIONS:
        has_images, image_count, has_merged_cells, sheet_count = _analyze_excel(file_path)
    elif extension == PDF_EXTENSION:
        has_images, image_count = _analyze_pdf(file_path, file_size)
    elif extension in OFFICE_EXTENSIONS:
        has_images, image_count = _analyze_office(file_path, extension)

    # Determine recommendation
    recommendation, reason = _make_recommendation(
        extension, has_images, image_count, has_merged_cells, file_size
    )

    return FileAnalysisResponse(
        filename=original_filename,
        extension=extension,
        file_size_bytes=file_size,
        has_images=has_images,
        image_count=image_count,
        has_merged_cells=has_merged_cells,
        sheet_count=sheet_count,
        recommendation=recommendation,
        recommendation_reason=reason,
    )


def _analyze_excel(file_path: Path) -> tuple[bool, int, bool, int]:
    """Analyze Excel file for images and merged cells."""
    has_images = False
    image_count = 0
    has_merged_cells = False
    sheet_count = 0

    try:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, data_only=True, read_only=False)
        sheet_count = len(wb.sheetnames)

        for ws in wb.worksheets:
            if hasattr(ws, "_images") and ws._images:
                has_images = True
                image_count += len(ws._images)
            if ws.merged_cells.ranges:
                has_merged_cells = True

        wb.close()
    except Exception:
        pass

    return has_images, image_count, has_merged_cells, sheet_count


def _analyze_pdf(file_path: Path, file_size: int) -> tuple[bool, int]:
    """Analyze PDF for embedded images."""
    has_images = False
    image_count = 0

    try:
        import pdfplumber
        with pdfplumber.open(str(file_path)) as pdf:
            for page in pdf.pages:
                if hasattr(page, "images") and page.images:
                    has_images = True
                    image_count += len(page.images)
    except ImportError:
        # pdfplumber not available - use heuristic based on file size vs text ratio
        # Large PDFs with little extractable text likely have images
        try:
            import pdfminer.high_level
            text = pdfminer.high_level.extract_text(str(file_path))
            text_ratio = len(text.strip()) / max(file_size, 1)
            if file_size > 100_000 and text_ratio < 0.05:
                has_images = True
                image_count = -1  # Unknown count
        except Exception:
            pass
    except Exception:
        pass

    return has_images, image_count


def _analyze_office(file_path: Path, extension: str) -> tuple[bool, int]:
    """Analyze DOCX/PPTX for images."""
    has_images = False
    image_count = 0

    try:
        import zipfile
        with zipfile.ZipFile(file_path, "r") as z:
            image_extensions = {".png", ".jpg", ".jpeg", ".gif", ".bmp", ".tiff", ".emf", ".wmf"}
            for name in z.namelist():
                if any(name.lower().endswith(ext) for ext in image_extensions):
                    # Skip thumbnail
                    if "thumbnail" not in name.lower():
                        has_images = True
                        image_count += 1
    except Exception:
        pass

    return has_images, image_count


def _make_recommendation(
    extension: str,
    has_images: bool,
    image_count: int,
    has_merged_cells: bool,
    file_size: int,
) -> tuple[str, str]:
    """Determine the best conversion mode recommendation."""

    # Text-based formats: always standard
    if extension in {".csv", ".json", ".xml", ".txt", ".md", ".html", ".htm"}:
        return "standard", "Text-based format — standard conversion provides high accuracy."

    # Excel with images
    if extension in EXCEL_EXTENSIONS:
        if has_images and image_count > 0:
            return (
                "ai_recommended",
                f"This Excel file contains {image_count} embedded image(s). "
                "Enable AI plugin to extract text and descriptions from images. "
                "Without AI, images will be silently dropped from the output."
            )
        if has_merged_cells:
            return (
                "standard",
                "Excel with merged cells detected. Standard conversion handles this, "
                "but review the output for any layout approximations."
            )
        return "standard", "Excel file with tabular data — standard conversion is sufficient."

    # PDF analysis
    if extension == PDF_EXTENSION:
        if has_images and image_count > 3:
            return (
                "ai_recommended",
                f"This PDF contains {image_count} image(s)/visual elements. "
                "Enable AI plugin for OCR + image description to capture all content. "
                "Without AI, images and scanned content will be missing from output."
            )
        if has_images:
            return (
                "ai_recommended",
                f"This PDF contains {image_count} image(s). "
                "AI plugin recommended for complete content extraction."
            )
        # Large PDF with no detected images but very little text = likely scanned
        if file_size > 200_000 and image_count == -1:
            return (
                "ai_required",
                "This PDF appears to be scanned/image-based. "
                "AI plugin with OCR is required to extract meaningful text."
            )
        return "standard", "Text-based PDF — standard conversion provides good accuracy."

    # DOCX/PPTX with images
    if extension in OFFICE_EXTENSIONS:
        if has_images and image_count > 0:
            return (
                "ai_recommended",
                f"This document contains {image_count} image(s). "
                "Enable AI plugin to generate descriptions of embedded images/charts. "
                "Without AI, images will be referenced but not described."
            )
        return "standard", "Document with text content — standard conversion is sufficient."

    return "standard", "Standard conversion will be used."
