"""Confidence scoring for markdown conversions.

Evaluates the quality and completeness of a conversion based on
known limitations, structural analysis, and content heuristics.
"""
from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.models import ConfidenceScore

EXCEL_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm"}


def compute_confidence(
    source_path: Path,
    markdown_output: str,
    *,
    extension: str,
    sheet_metadata: list[dict[str, Any]] | None = None,
    plugin_enabled: bool = False,
    images_found: int = 0,
    images_described: int = 0,
) -> ConfidenceScore:
    """Compute a confidence score for the conversion."""
    factors: list[str] = []
    limitations: list[str] = []

    structure_score = 100.0
    content_score = 100.0
    image_score = 100.0
    formatting_score = 100.0

    # --- Excel-specific analysis ---
    if extension in EXCEL_EXTENSIONS:
        structure_score, content_score, image_score, formatting_score = _score_excel(
            source_path, markdown_output, sheet_metadata or [], factors, limitations, plugin_enabled
        )
    # --- PDF analysis ---
    elif extension == ".pdf":
        structure_score, content_score, image_score, formatting_score = _score_pdf(
            source_path, markdown_output, factors, limitations, plugin_enabled
        )
    # --- Other document types ---
    elif extension in {".docx", ".doc", ".pptx"}:
        structure_score, content_score, image_score, formatting_score = _score_office_doc(
            extension, markdown_output, factors, limitations, plugin_enabled
        )
    # --- Simple text formats ---
    elif extension in {".csv", ".json", ".xml", ".txt", ".md", ".html", ".htm"}:
        factors.append("Text-based format: high confidence in content extraction")
        structure_score = 95.0
        content_score = 98.0
        image_score = 100.0  # no images expected
        formatting_score = 90.0
    else:
        factors.append(f"Extension '{extension}' uses auto-detection; confidence may vary")
        structure_score = 60.0
        content_score = 70.0

    # --- Output emptiness check ---
    if not markdown_output.strip():
        content_score = 0.0
        factors.append("Output is empty - conversion may have failed")
        limitations.append("Empty output: file may require OCR or specialized parser")

    # --- Markdown quality heuristics ---
    if markdown_output.strip():
        table_count = markdown_output.count("| ---")
        if extension in EXCEL_EXTENSIONS and table_count == 0:
            structure_score *= 0.5
            factors.append("No markdown tables found in Excel conversion")

        # Check for NaN remnants
        nan_count = len(re.findall(r"\bNaN\b", markdown_output))
        if nan_count > 0:
            content_score -= min(nan_count * 2, 20)
            factors.append(f"Found {nan_count} NaN values (blank cells not cleaned)")

    # --- Image analysis quality (overrides earlier image_score if AI was used) ---
    if images_found > 0:
        if images_described > 0:
            coverage = images_described / images_found
            image_score = 60.0 + (coverage * 40.0)  # 60-100 based on coverage
            factors.append(
                f"AI analyzed {images_described}/{images_found} images "
                f"(OCR + description embedded inline)"
            )
            if coverage >= 1.0:
                factors.append("All embedded images fully described with OCR and AI")
        elif plugin_enabled:
            image_score = 30.0
            factors.append(f"Found {images_found} images but AI analysis returned no content")
            limitations.append("Images detected but AI could not extract meaningful content")
        else:
            image_score = 0.0
            limitations.append(
                f"Document contains {images_found} image(s) silently dropped "
                "(enable AI plugin for OCR + description)"
            )
    elif "IMAGE_START" in markdown_output or "Image OCR" in markdown_output:
        # Images were described inline (from MarkItDown llm_client)
        image_score = 85.0
        factors.append("Images described via AI plugin (inline)")

    overall = (
        structure_score * 0.35
        + content_score * 0.35
        + image_score * 0.15
        + formatting_score * 0.15
    )

    return ConfidenceScore(
        overall=round(min(max(overall, 0), 100), 1),
        structure_fidelity=round(min(max(structure_score, 0), 100), 1),
        content_completeness=round(min(max(content_score, 0), 100), 1),
        image_handling=round(min(max(image_score, 0), 100), 1),
        formatting_preservation=round(min(max(formatting_score, 0), 100), 1),
        factors=factors,
        limitations_hit=limitations,
    )


def _score_excel(
    source_path: Path,
    markdown_output: str,
    sheet_metadata: list[dict[str, Any]],
    factors: list[str],
    limitations: list[str],
    plugin_enabled: bool,
) -> tuple[float, float, float, float]:
    structure_score = 90.0
    content_score = 90.0
    image_score = 100.0
    formatting_score = 80.0

    try:
        wb = load_workbook(source_path, data_only=True, read_only=False)

        # Check for embedded images
        has_images = False
        for ws in wb.worksheets:
            if ws._images:
                has_images = True
                break

        if has_images:
            if plugin_enabled:
                image_score = 75.0
                factors.append("Excel contains images: AI plugin used for description")
            else:
                image_score = 0.0
                limitations.append("Excel contains embedded images that are NOT extracted (enable AI plugin for image descriptions)")
                factors.append("Images in Excel are silently dropped without plugin")

        # Check for merged cells
        has_merged = False
        for ws in wb.worksheets:
            if ws.merged_cells.ranges:
                has_merged = True
                break

        if has_merged:
            structure_score -= 15.0
            factors.append("Merged cells detected: structure may be approximated")
            limitations.append("Merged cells are expanded but original layout context may be lost")

        # Check for formulas (data_only=True returns cached values)
        # If a formula cell has no cached value, it shows as None
        empty_formula_cells = 0
        total_cells = 0
        for ws in wb.worksheets:
            for row in ws.iter_rows(max_row=min(ws.max_row or 1, 100)):
                for cell in row:
                    total_cells += 1
                    if cell.value is None and cell.data_type == "f":
                        empty_formula_cells += 1

        if empty_formula_cells > 0:
            content_score -= min(empty_formula_cells * 5, 25)
            factors.append(f"{empty_formula_cells} formula cells with no cached value")
            limitations.append("Some formula results not available (file may need to be opened in Excel first)")

        # Check sheet count vs converted
        total_sheets = len(wb.sheetnames)
        converted_sheets = sum(1 for s in sheet_metadata if s.get("status") == "converted")
        if converted_sheets < total_sheets:
            content_score -= (total_sheets - converted_sheets) * 10
            factors.append(f"Only {converted_sheets}/{total_sheets} sheets had convertible content")

        # Wide table penalty
        for s in sheet_metadata:
            cols = s.get("column_count", 0)
            if cols > 40:
                structure_score -= 10
                limitations.append(f"Sheet '{s.get('sheet_name')}' has {cols} columns: may exceed LLM context window")

        wb.close()
    except Exception:
        structure_score = 60.0
        factors.append("Could not re-analyze source Excel for scoring")

    # Currency / date formatting is lost
    formatting_score = 70.0
    limitations.append("Cell formatting (currency symbols, date formats, colors, conditional formatting) is not preserved")

    return structure_score, content_score, image_score, formatting_score


def _score_pdf(
    source_path: Path,
    markdown_output: str,
    factors: list[str],
    limitations: list[str],
    plugin_enabled: bool,
) -> tuple[float, float, float, float]:
    structure_score = 75.0
    content_score = 80.0
    image_score = 100.0
    formatting_score = 65.0

    file_size = source_path.stat().st_size
    output_len = len(markdown_output.strip())

    # Heuristic: if file is large but output is tiny, likely scanned/image PDF
    if file_size > 100_000 and output_len < 200:
        if plugin_enabled:
            factors.append("PDF appears to be scanned/image-heavy: AI plugin may help with OCR")
            image_score = 50.0
            content_score = 50.0
        else:
            image_score = 0.0
            content_score = 20.0
            factors.append("PDF appears to be scanned/image-based with very little extracted text")
            limitations.append("Scanned/image PDFs require AI plugin (OCR) to extract text from images")
    else:
        factors.append("PDF text extraction via built-in converter")

    # Check for table structures
    if "| ---" in markdown_output:
        factors.append("Tables detected in PDF output")
        structure_score = 70.0
        limitations.append("PDF table extraction may misalign columns in complex layouts")
    else:
        if file_size > 50_000:
            limitations.append("PDF tables (if any) may not be perfectly structured in output")

    # Images in PDF
    # We can't easily detect images without parsing, use size heuristic
    if file_size > 500_000 and output_len < file_size * 0.01:
        if not plugin_enabled:
            image_score = 20.0
            limitations.append("PDF likely contains images/charts not represented in text output (enable AI plugin)")
        else:
            image_score = 60.0
            factors.append("AI plugin enabled for image/OCR processing")

    formatting_score = 60.0
    limitations.append("PDF formatting (fonts, colors, multi-column layouts) is not preserved")

    return structure_score, content_score, image_score, formatting_score


def _score_office_doc(
    extension: str,
    markdown_output: str,
    factors: list[str],
    limitations: list[str],
    plugin_enabled: bool,
) -> tuple[float, float, float, float]:
    structure_score = 80.0
    content_score = 85.0
    image_score = 100.0
    formatting_score = 70.0

    # Check for image references in output
    has_image_refs = bool(re.search(r"!\[.*?\]\(.*?\)", markdown_output))
    has_base64_refs = "data:image" in markdown_output

    if has_image_refs or has_base64_refs:
        if plugin_enabled:
            image_score = 80.0
            factors.append("Document images detected; AI plugin provides descriptions")
        else:
            image_score = 40.0
            factors.append("Document contains images (referenced but not described)")
            limitations.append("Images are referenced but not described without AI plugin")
    else:
        if extension == ".pptx":
            # PPTX often has images
            image_score = 50.0
            limitations.append("PowerPoint slides may contain images/charts not fully captured")
        factors.append(f"{extension.upper()} text and structure extracted")

    if extension == ".pptx":
        limitations.append("Slide animations, transitions, and speaker notes layout may not be preserved")
    elif extension in {".docx", ".doc"}:
        limitations.append("Track changes, comments, and complex formatting are not preserved")

    return structure_score, content_score, image_score, formatting_score
