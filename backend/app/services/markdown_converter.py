from __future__ import annotations

import hashlib
import json
import re
import shutil
import sys
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell, MergedCell

from app.config import (
    MAX_UPLOAD_BYTES,
    MANIFEST_DIR,
    MARKDOWN_OUTPUT_DIR,
    MARKITDOWN_SOURCE_DIR,
    RAW_UPLOAD_DIR,
    ensure_data_dirs,
)
from app.models import ConversionMetadata, ConversionResponse, ConfidenceScore, AICostEstimate
from app.services.confidence_scorer import compute_confidence
from app.services.ai_plugin import (
    convert_with_plugin,
    is_plugin_available,
    analyze_document_images,
    format_image_block,
)

if MARKITDOWN_SOURCE_DIR.exists() and str(MARKITDOWN_SOURCE_DIR) not in sys.path:
    sys.path.insert(0, str(MARKITDOWN_SOURCE_DIR))

try:
    from markitdown import MarkItDown
except ImportError:  # pragma: no cover - surfaced in API response
    MarkItDown = None  # type: ignore[assignment]


EXCEL_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm"}
MARKITDOWN_EXTENSIONS = {
    ".csv",
    ".doc",
    ".docx",
    ".epub",
    ".html",
    ".htm",
    ".json",
    ".md",
    ".msg",
    ".pdf",
    ".xls",
    ".pptx",
    ".rtf",
    ".txt",
    ".xml",
    ".zip",
}


@dataclass
class SheetConversion:
    name: str
    markdown: str
    metadata: dict[str, Any]
    warnings: list[str]


def convert_upload(
    source_file: Path,
    *,
    original_filename: str,
    document_title: str | None = None,
    department: str | None = None,
    access_group: str | None = None,
    classification: str = "Internal",
    document_owner: str | None = None,
    additional_context: str | None = None,
    version: str | None = None,
    expiry_review_date: str | None = None,
    enable_plugin: bool = False,
) -> ConversionResponse:
    ensure_data_dirs()

    if not source_file.exists():
        raise ValueError(f"Source file does not exist: {source_file}")
    if source_file.stat().st_size == 0:
        raise ValueError("Uploaded file is empty.")
    if source_file.stat().st_size > MAX_UPLOAD_BYTES:
        raise ValueError(f"Uploaded file exceeds the {MAX_UPLOAD_BYTES // (1024 * 1024)} MB limit.")

    original_filename = _safe_original_filename(original_filename)
    file_hash = _sha256(source_file)
    extension = source_file.suffix.lower()
    safe_stem = _slugify(Path(original_filename).stem) or file_hash[:12]
    document_id = f"{safe_stem}-{file_hash[:12]}"
    saved_source = RAW_UPLOAD_DIR / f"{document_id}{extension}"
    shutil.copy2(source_file, saved_source)

    title = document_title.strip() if document_title and document_title.strip() else Path(original_filename).stem
    converted_at = datetime.now(UTC).replace(microsecond=0).isoformat()
    metadata = ConversionMetadata(
        document_id=document_id,
        document_title=title,
        source_filename=original_filename,
        source_type=extension.removeprefix(".").upper() or "UNKNOWN",
        source_path=str(saved_source),
        markdown_path="",
        file_sha256=file_hash,
        file_size_bytes=source_file.stat().st_size,
        converted_at_utc=converted_at,
        converter="knowledge-hub-markdown-converter",
        extraction_strategy="",
        department=_empty_to_none(department),
        access_group=_empty_to_none(access_group),
        classification=classification.strip() or "Internal",
        document_owner=_empty_to_none(document_owner),
        additional_context=_empty_to_none(additional_context),
        version=_empty_to_none(version),
        expiry_review_date=_empty_to_none(expiry_review_date),
    )

    warnings: list[str] = []
    plugin_used = False
    plugin_name: str | None = None
    sheet_metadata: list[dict[str, Any]] = []
    image_count = 0
    images_described = 0
    ai_cost_estimates: list[AICostEstimate] = []

    if extension in EXCEL_EXTENSIONS:
        markdown_body, sheet_metadata, sheet_warnings = _convert_excel(saved_source)
        metadata.extraction_strategy = "openpyxl workbook extraction with sheet-level Markdown tables"
        metadata.sheet_count = len(sheet_metadata)
        metadata.sheets = sheet_metadata
        warnings.extend(sheet_warnings)

        # If plugin enabled, extract and analyze embedded images
        if enable_plugin and is_plugin_available():
            try:
                image_analyses, img_cost = analyze_document_images(saved_source, extension)
                image_count = len(image_analyses)
                if image_analyses:
                    image_section_parts = ["\n\n## Embedded Images\n"]
                    image_section_parts.append(
                        f"_This workbook contained {len(image_analyses)} embedded image(s). "
                        "AI analysis below:_\n"
                    )
                    for analysis in image_analyses:
                        image_section_parts.append(format_image_block(analysis))
                        if analysis.description or analysis.ocr_text:
                            images_described += 1
                    markdown_body += "\n".join(image_section_parts)
                    metadata.extraction_strategy = "openpyxl + Azure OpenAI (image OCR + description)"
                    plugin_used = True
                    plugin_name = "Azure OpenAI (GPT-4o)"
                if img_cost:
                    ai_cost_estimates.append(img_cost)
            except Exception as e:
                warnings.append(f"AI image analysis attempted but failed: {e}")

    elif extension in MARKITDOWN_EXTENSIONS:
        if enable_plugin and is_plugin_available() and extension in {".pdf", ".docx", ".pptx"}:
            try:
                # First: get the standard text conversion
                markdown_body, plugin_cost = convert_with_plugin(saved_source)
                metadata.extraction_strategy = "MarkItDown + Azure OpenAI plugin (image-aware)"
                plugin_used = True
                plugin_name = "Azure OpenAI (GPT-4o)"
                if plugin_cost:
                    ai_cost_estimates.append(plugin_cost)

                # Second: also extract and analyze images separately for inline placement
                if extension == ".pdf":
                    try:
                        image_analyses, img_cost = analyze_document_images(saved_source, extension)
                        image_count = len(image_analyses)
                        if image_analyses:
                            image_section_parts = ["\n\n## Embedded Images & Visual Content\n"]
                            image_section_parts.append(
                                f"_This document contained {len(image_analyses)} image(s). "
                                "AI OCR + description below:_\n"
                            )
                            for analysis in image_analyses:
                                image_section_parts.append(format_image_block(analysis))
                                if analysis.description or analysis.ocr_text:
                                    images_described += 1
                            markdown_body += "\n".join(image_section_parts)
                            metadata.extraction_strategy = "MarkItDown + Azure OpenAI (text + image OCR + description)"
                        if img_cost:
                            ai_cost_estimates.append(img_cost)
                    except Exception as e:
                        warnings.append(f"PDF image extraction partial failure: {e}")
            except Exception as e:
                warnings.append(f"AI plugin failed, falling back to standard: {e}")
                markdown_body = _convert_with_markitdown(saved_source)
                metadata.extraction_strategy = "MarkItDown built-in converter"
        else:
            markdown_body = _convert_with_markitdown(saved_source)
            metadata.extraction_strategy = "MarkItDown built-in converter"
    else:
        markdown_body = _convert_with_markitdown(saved_source)
        metadata.extraction_strategy = "MarkItDown auto-detected converter"
        warnings.append(f"The extension '{extension or 'unknown'}' is not explicitly listed; conversion used auto-detection.")

    if not markdown_body.strip():
        warnings.append("Conversion completed, but the Markdown body is empty. This file may need OCR or a custom parser.")

    # Report image handling results
    if image_count > 0 and plugin_used:
        warnings.append(
            f"Found {image_count} embedded image(s): {images_described} successfully analyzed with AI."
        )

    # Compute confidence score
    confidence = compute_confidence(
        saved_source,
        markdown_body,
        extension=extension,
        sheet_metadata=sheet_metadata,
        plugin_enabled=plugin_used,
        images_found=image_count,
        images_described=images_described,
    )

    output_path = MARKDOWN_OUTPUT_DIR / f"{document_id}.md"
    manifest_path = MANIFEST_DIR / f"{document_id}.json"
    metadata.markdown_path = str(output_path)

    markdown = _build_markdown(metadata, markdown_body)
    output_path.write_text(markdown, encoding="utf-8")
    manifest_path.write_text(
        json.dumps(metadata.model_dump(), indent=2, ensure_ascii=False),
        encoding="utf-8",
    )

    # Aggregate AI cost estimates from all plugin calls
    ai_cost: AICostEstimate | None = None
    if ai_cost_estimates:
        if len(ai_cost_estimates) == 1:
            ai_cost = ai_cost_estimates[0]
        else:
            # Merge all estimates into one
            from app.models import AICostCall
            all_calls: list[AICostCall] = []
            total_in = 0
            total_out = 0
            model_name = ai_cost_estimates[0].model
            for est in ai_cost_estimates:
                all_calls.extend(est.cost_per_call)
                total_in += est.total_input_tokens
                total_out += est.total_output_tokens
            from app.services.ai_plugin import _get_pricing
            inp_per_1m, out_per_1m = _get_pricing(model_name)
            total_cost = total_in / 1_000_000 * inp_per_1m + total_out / 1_000_000 * out_per_1m
            ai_cost = AICostEstimate(
                model=model_name,
                total_input_tokens=total_in,
                total_output_tokens=total_out,
                total_cost_usd=round(total_cost, 6),
                cost_per_call=all_calls,
            )

    return ConversionResponse(
        markdown=markdown,
        metadata=metadata,
        output_file=str(output_path),
        manifest_file=str(manifest_path),
        warnings=warnings,
        confidence=confidence,
        plugin_used=plugin_used,
        plugin_name=plugin_name,
        ai_cost=ai_cost,
    )


def _convert_excel(path: Path) -> tuple[str, list[dict[str, Any]], list[str]]:
    workbook = load_workbook(path, data_only=True, read_only=False)
    sections: list[str] = []
    workbook_metadata: list[dict[str, Any]] = []
    warnings: list[str] = []
    workbook_title = path.stem

    for worksheet in workbook.worksheets:
        if worksheet.sheet_state != "visible":
            warnings.append(
                f"Sheet '{worksheet.title}' is {worksheet.sheet_state}; verify it should be indexed before embedding."
            )

        grid = _worksheet_to_grid(worksheet)
        trimmed_grid, row_offset, col_offset = _trim_grid(grid)
        if not trimmed_grid:
            workbook_metadata.append(
                {
                    "sheet_name": worksheet.title,
                    "status": "empty",
                    "row_count": 0,
                    "column_count": 0,
                }
            )
            warnings.append(f"Sheet '{worksheet.title}' was empty after trimming.")
            continue

        header_index = _infer_header_index(trimmed_grid)
        header = _make_headers(trimmed_grid[header_index])
        pre_header_rows = trimmed_grid[:header_index]
        data_rows = _clean_data_rows(trimmed_grid[header_index + 1 :], len(header))
        source_header_row = row_offset + header_index + 1
        source_start_col = col_offset + 1

        if len(header) > 40:
            warnings.append(
                f"Sheet '{worksheet.title}' has {len(header)} columns; row-group chunking is recommended before embedding."
            )
        if len(data_rows) > 500:
            warnings.append(
                f"Sheet '{worksheet.title}' has {len(data_rows)} data rows; split into smaller row chunks before embedding."
            )
        if _has_multiple_table_blocks(trimmed_grid, header_index):
            warnings.append(
                f"Sheet '{worksheet.title}' appears to contain multiple table blocks separated by blank rows; verify the inferred header."
            )

        sheet_section = [f"## Sheet: {_escape_markdown_text(worksheet.title)}"]
        sheet_section.append("")
        sheet_section.append(
            f"- Source rows: {row_offset + 1}-{row_offset + len(trimmed_grid)}"
        )
        sheet_section.append(
            f"- Source columns: {source_start_col}-{col_offset + len(trimmed_grid[0])}"
        )
        sheet_section.append(f"- Header row: {source_header_row}")

        notes = _rows_to_notes(pre_header_rows)
        if notes:
            sheet_section.append("")
            sheet_section.append("### Sheet Notes")
            sheet_section.extend(f"- {note}" for note in notes)

        sheet_section.append("")

        # Determine best output format for vector store consumption
        output_mode = _choose_output_mode(header, data_rows)

        if output_mode == "natural_language":
            # Natural language rows — each row becomes a self-contained statement
            # Best for wide pivot tables (Plan × ProcCode → Cover) common in health insurance data
            sheet_section.append("### Records (AI-Optimized Format)")
            sheet_section.append("")
            sheet_section.append(
                f"_Each record below is a self-contained fact from workbook "
                f"'{_escape_markdown_text(workbook_title)}', sheet '{_escape_markdown_text(worksheet.title)}'._"
            )
            sheet_section.append("")
            nl_rows = _to_natural_language_rows(header, data_rows)
            # Insert chunk boundary hints every 15 records for vector store indexers
            for i, row_text in enumerate(nl_rows):
                sheet_section.append(row_text)
                if (i + 1) % 15 == 0 and i < len(nl_rows) - 1:
                    sheet_section.append("")
                    sheet_section.append("---")
                    sheet_section.append("")
        elif output_mode == "key_value":
            # Key-value pairs — for sheets with 1-3 data rows (lookup/definitions)
            sheet_section.append("### Key-Value Pairs")
            sheet_section.append("")
            for row in data_rows:
                pairs = [f"**{header[j]}**: {row[j]}" for j in range(len(header)) if row[j]]
                if pairs:
                    sheet_section.append("- " + " | ".join(pairs))
        else:
            # Standard table output with context header for chunking
            sheet_section.append("### Table")
            sheet_section.append("")
            sheet_section.append(
                f"_Source: {_escape_markdown_text(workbook_title)} > "
                f"{_escape_markdown_text(worksheet.title)} "
                f"| Headers: {', '.join(header[:8])}"
                f"{'...' if len(header) > 8 else ''}_"
            )
            sheet_section.append("")
            if data_rows:
                # For large tables, insert chunk boundary hints every 25 rows
                if len(data_rows) > 25:
                    chunks = _chunk_rows(data_rows, 25)
                    for chunk_idx, chunk in enumerate(chunks):
                        if chunk_idx > 0:
                            sheet_section.append("")
                            sheet_section.append("---")
                            sheet_section.append("")
                            sheet_section.append(
                                f"_Continued: {_escape_markdown_text(workbook_title)} > "
                                f"{_escape_markdown_text(worksheet.title)} "
                                f"(rows {chunk_idx * 25 + 1}-{chunk_idx * 25 + len(chunk)})_"
                            )
                            sheet_section.append("")
                        sheet_section.append(_markdown_table(header, chunk))
                else:
                    sheet_section.append(_markdown_table(header, data_rows))
            else:
                sheet_section.append("No data rows were found after the inferred header row.")

        sections.append("\n".join(sheet_section).strip())
        workbook_metadata.append(
            {
                "sheet_name": worksheet.title,
                "status": "converted",
                "sheet_state": worksheet.sheet_state,
                "row_count": len(data_rows),
                "column_count": len(header),
                "header_row": source_header_row,
                "source_row_start": row_offset + 1,
                "source_row_end": row_offset + len(trimmed_grid),
                "source_column_start": source_start_col,
                "source_column_end": col_offset + len(trimmed_grid[0]),
                "headers": header,
                "output_mode": output_mode,
            }
        )

    workbook.close()
    return "\n\n".join(sections).strip(), workbook_metadata, warnings


def _choose_output_mode(header: list[str], data_rows: list[list[str]]) -> str:
    """Decide the best output format for AI/vector store ingestion.

    Returns:
        "natural_language" — wide pivot tables where most cells repeat a small set of values
        "key_value" — very few rows (definitions, lookups)
        "table" — standard Markdown table (default)
    """
    if len(data_rows) <= 3:
        return "key_value"

    # Detect pivot-like patterns: many columns (>15), most cells from a small vocabulary
    if len(header) > 15 and len(data_rows) > 5:
        # Sample data cells (excluding first 2 cols which are typically row identifiers)
        sample_values: list[str] = []
        for row in data_rows[:50]:
            sample_values.extend(cell.strip().lower() for cell in row[2:] if cell.strip())
        if sample_values:
            unique_values = set(sample_values)
            # If 80%+ of cells come from ≤10 distinct values → it's a pivot
            if len(unique_values) <= 10 and len(sample_values) > 20:
                return "natural_language"

    # Wide tables (>30 cols) are always better as natural language for embeddings
    if len(header) > 30:
        return "natural_language"

    return "table"


def _to_natural_language_rows(header: list[str], data_rows: list[list[str]]) -> list[str]:
    """Convert wide pivot-style rows into self-contained natural language statements.

    For tables like: [PlanCode | PlanName | Proc1 | Proc2 | ... | Proc40]
    Produces: "Plan E74 (360 Care): Proc1=Full cover, Proc2=Full cover, ..."

    This format embeds MUCH better than 40-column tables because each fact is
    semantically complete and can be matched by vector similarity search.
    """
    results: list[str] = []

    # Detect identifier columns: the leading columns whose name suggests an
    # identifier (code, plan, id, label, no., description), plus the first two
    # columns unconditionally, capped at three.
    id_patterns = re.compile(r"(code|plan|id|label|no\.|number|scheme|description|name|column 1)", re.IGNORECASE)
    id_cols: list[int] = []
    for i in range(min(5, len(header))):
        if id_patterns.search(header[i]) or i < 2:
            id_cols.append(i)
        if len(id_cols) >= 3:
            break
    if not id_cols:
        id_cols = [0]

    value_cols = [i for i in range(len(header)) if i not in id_cols]

    for row in data_rows:
        # Build identifier part
        id_parts = [f"{header[i]}: {row[i]}" for i in id_cols if row[i]]
        if not id_parts:
            continue
        id_str = " | ".join(id_parts)

        # Build value part — group by value to compress
        value_groups: dict[str, list[str]] = {}
        for col_idx in value_cols:
            cell_val = row[col_idx].strip() if col_idx < len(row) else ""
            if cell_val:
                value_groups.setdefault(cell_val, []).append(header[col_idx])

        if not value_groups:
            continue

        # Format compactly: group same-value columns
        value_parts: list[str] = []
        for val, cols in value_groups.items():
            if len(cols) <= 3:
                value_parts.append(f"{', '.join(cols)} = {val}")
            else:
                value_parts.append(f"{val} (×{len(cols)} procedures: {', '.join(cols[:3])}...)")

        row_text = f"- **{id_str}** → {'; '.join(value_parts)}"
        results.append(row_text)

    return results


def _chunk_rows(rows: list[list[str]], chunk_size: int) -> list[list[list[str]]]:
    """Split rows into chunks for vector store boundary hints."""
    return [rows[i:i + chunk_size] for i in range(0, len(rows), chunk_size)]


def _worksheet_to_grid(worksheet: Any) -> list[list[str]]:
    merged_values: dict[tuple[int, int], Any] = {}
    for merged_range in worksheet.merged_cells.ranges:
        anchor_value = worksheet.cell(merged_range.min_row, merged_range.min_col).value
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for column in range(merged_range.min_col, merged_range.max_col + 1):
                merged_values[(row, column)] = anchor_value

    grid: list[list[str]] = []
    for row in worksheet.iter_rows():
        values: list[str] = []
        for cell in row:
            value = merged_values.get((cell.row, cell.column), cell.value)
            values.append(_normalise_cell(value, cell))
        grid.append(values)
    return grid


def _trim_grid(grid: list[list[str]]) -> tuple[list[list[str]], int, int]:
    non_empty_rows = [index for index, row in enumerate(grid) if any(row)]
    if not non_empty_rows:
        return [], 0, 0

    first_row = non_empty_rows[0]
    last_row = non_empty_rows[-1]
    candidate_rows = grid[first_row : last_row + 1]
    non_empty_columns = [
        index
        for index in range(max(len(row) for row in candidate_rows))
        if any(index < len(row) and row[index] for row in candidate_rows)
    ]
    first_col = non_empty_columns[0]
    last_col = non_empty_columns[-1]
    trimmed = [row[first_col : last_col + 1] for row in candidate_rows]
    width = max(len(row) for row in trimmed)
    return [row + [""] * (width - len(row)) for row in trimmed], first_row, first_col


def _infer_header_index(grid: list[list[str]]) -> int:
    search_window = grid[: min(len(grid), 25)]
    best_index = 0
    best_score = -1.0
    for index, row in enumerate(search_window):
        cells = [cell for cell in row if cell]
        if len(cells) < 2:
            continue
        text_cells = sum(1 for cell in cells if re.search(r"[A-Za-z]", cell))
        unique_cells = len({cell.lower() for cell in cells})
        numeric_cells = sum(1 for cell in cells if re.fullmatch(r"[-+]?\d+(\.\d+)?", cell))
        score = (len(cells) * 2) + text_cells + (unique_cells * 0.5) - (numeric_cells * 0.25) - (index * 0.05)
        if score > best_score:
            best_score = score
            best_index = index
    return best_index


def _make_headers(row: list[str]) -> list[str]:
    headers: list[str] = []
    seen: dict[str, int] = {}
    for index, value in enumerate(row, start=1):
        header = value.strip() if value else f"Column {index}"
        header = re.sub(r"\s+", " ", header)
        if header.lower().startswith("unnamed"):
            header = f"Column {index}"
        count = seen.get(header.lower(), 0)
        seen[header.lower()] = count + 1
        if count:
            header = f"{header} {count + 1}"
        headers.append(header)
    return headers


def _clean_data_rows(rows: list[list[str]], width: int) -> list[list[str]]:
    cleaned: list[list[str]] = []
    for row in rows:
        normalised = (row + [""] * width)[:width]
        if any(normalised):
            cleaned.append(normalised)
    return cleaned


def _has_multiple_table_blocks(grid: list[list[str]], header_index: int) -> bool:
    seen_data = False
    seen_blank_after_data = False
    for row in grid[header_index + 1 :]:
        if any(row):
            if seen_blank_after_data:
                return True
            seen_data = True
        elif seen_data:
            seen_blank_after_data = True
    return False


def _rows_to_notes(rows: list[list[str]]) -> list[str]:
    notes: list[str] = []
    for row in rows:
        values = [value for value in row if value]
        if values:
            notes.append(" | ".join(_escape_markdown_text(value) for value in values))
    return notes


def _markdown_table(headers: list[str], rows: list[list[str]]) -> str:
    lines = [
        "| " + " | ".join(_escape_table_cell(header) for header in headers) + " |",
        "| " + " | ".join("---" for _ in headers) + " |",
    ]
    for row in rows:
        lines.append("| " + " | ".join(_escape_table_cell(cell) for cell in row) + " |")
    return "\n".join(lines)


def _convert_with_markitdown(path: Path) -> str:
    if MarkItDown is None:
        raise RuntimeError(
            "MarkItDown is not available. Install the backend requirements with "
            "`pip install -r requirements.txt`."
        )
    result = MarkItDown().convert(str(path))
    return result.text_content.strip()


def _build_markdown(metadata: ConversionMetadata, body: str) -> str:
    frontmatter = _yaml_frontmatter(metadata.model_dump(exclude={"markdown_path"}))
    title = _escape_markdown_text(metadata.document_title)
    context_section = _additional_context_section(metadata.additional_context)
    return f"---\n{frontmatter}---\n\n# {title}\n\n{context_section}{body.strip()}\n"


def _additional_context_section(additional_context: str | None) -> str:
    if not additional_context:
        return ""
    normalized_context = re.sub(r"\n{3,}", "\n\n", additional_context.strip())
    return f"## Additional Context Provided by User\n\n{normalized_context}\n\n"


def _yaml_frontmatter(data: dict[str, Any]) -> str:
    lines: list[str] = []
    for key, value in data.items():
        if value is None or value == []:
            continue
        if isinstance(value, list):
            lines.append(f"{key}:")
            for item in value:
                lines.append("  - " + json.dumps(item, ensure_ascii=False))
        else:
            lines.append(f"{key}: {json.dumps(value, ensure_ascii=False)}")
    return "\n".join(lines) + "\n"


def _normalise_cell(value: Any, cell: Cell | MergedCell) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat() if value.time().isoformat() == "00:00:00" else value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).replace("\u00a0", " ").strip()
    text = re.sub(r"\s+", " ", text)
    if text.lower() == "nan":
        return ""
    return text


def _escape_table_cell(value: str) -> str:
    return _escape_markdown_text(value).replace("\n", "<br>")


def _escape_markdown_text(value: str) -> str:
    return value.replace("|", "\\|").strip()


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _slugify(value: str) -> str:
    value = value.lower().strip()
    value = re.sub(r"[^a-z0-9]+", "-", value)
    return value.strip("-")[:80]


def _safe_original_filename(value: str) -> str:
    filename = Path(value).name.replace("\x00", "").strip()
    return filename or "upload"


def _empty_to_none(value: str | None) -> str | None:
    if value is None:
        return None
    value = value.strip()
    return value or None